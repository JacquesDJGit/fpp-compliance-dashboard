from __future__ import annotations

from datetime import datetime
from html import escape
from pathlib import Path
import re

from openpyxl import Workbook, load_workbook


BASE_DIR = Path(__file__).parent
TRACKER_PATH = BASE_DIR / "compliance_tracker.xlsx"
OUTPUT_HTML = BASE_DIR / "index.html"


URL_RE = re.compile(r"(https?://[^\s]+)")


def linkify(text: str) -> str:
    if not text:
        return ""
    chunks: list[str] = []
    last = 0
    for match in URL_RE.finditer(text):
        start, end = match.span()
        chunks.append(escape(text[last:start]))
        url = match.group(1)
        chunks.append(
            f'<a href="{escape(url)}" target="_blank">{escape(url.replace("https://", "").replace("http://", ""))}</a>'
        )
        last = end
    chunks.append(escape(text[last:]))
    return "".join(chunks)


def status_class(value: str) -> str:
    lookup = {
        "complete": "s-complete",
        "pending": "s-pending",
        "overdue": "s-overdue",
        "upcoming": "s-upcoming",
        "na": "s-na",
        "no": "s-na",
    }
    return lookup.get((value or "").strip().lower(), "s-upcoming")


def auth_class(value: str) -> str:
    key = (value or "").strip().lower()
    if "sars" in key:
        return "auth-sars"
    if "cipc" in key:
        return "auth-cipc"
    if "popia" in key:
        return "auth-popia"
    return "auth-sha"


def create_default_tracker(path: Path) -> None:
    wb = Workbook()

    meta = wb.active
    meta.title = "Meta"
    meta.append(["key", "value"])
    meta.append(["title", "FPP Compliance Obligations"])
    meta.append(["company", "Five Peaks Properties Proprietary Limited"])
    meta.append(["registration", "2025/581417/07"])
    meta.append(["year_end", "30 June"])
    meta.append(["as_at", "7 April 2026"])
    meta.append(
        [
            "summary",
            "FPP is a South African private property company incorporated in July 2025, operating one short-term rental property with a mortgage in place and no employees. The rotating shareholder handles bookkeeping and secretarial duties. All compliance obligations are listed below.",
        ]
    )
    meta.append(
        [
            "footer",
            "Prepared based on Five Peaks Properties Shareholders' Agreement dated 22 July 2025 and the South African compliance framework for private companies. This document does not constitute legal or tax advice.",
        ]
    )

    once_off = wb.create_sheet("OnceOff")
    once_off.append(["obligation", "authority", "status", "notes"])
    once_off.append(
        [
            "Register on SARS eFiling",
            "SARS",
            "Pending",
            "Tax reference number exists. eFiling access needed to file IRP6 and ITR14. Register at https://www.sars.gov.za/efiling",
        ]
    )
    once_off.append(
        [
            "Appoint Information Officer (POPIA)",
            "POPIA",
            "Pending",
            "Register with the Information Regulator. Urgent given managing agent handles tenant personal data. Register at https://www.inforegulator.org.za",
        ]
    )
    once_off.append(
        [
            "Beneficial Ownership Register (CIPC)",
            "CIPC",
            "Complete",
            "Filed July 2025. Equal 20% shareholding across all five entities.",
        ]
    )

    annual = wb.create_sheet("Annual")
    annual.append(["obligation", "due_date", "authority", "status", "notes", "overdue_row"])
    annual.append(
        [
            "1st Provisional Tax - IRP6",
            "31 Dec 2025",
            "SARS",
            "Overdue",
            "Must file IRP6 return on eFiling before payment. Estimate taxable income for FY ending 30 Jun 2026. Deductibles: bond interest, agent fees, rates, insurance. Late filing risks 20% underestimation penalty plus interest.",
            "Y",
        ]
    )
    annual.append(
        [
            "Annual Financial Statements",
            "30 Jun 2026",
            "SHA §15",
            "Upcoming",
            "Rotating shareholder responsible. No audit required unless majority votes for one. Required to support ITR14 filing.",
            "",
        ]
    )
    annual.append(
        [
            "2nd Provisional Tax - IRP6",
            "30 Jun 2026",
            "SARS",
            "Upcoming",
            "Based on actual FY figures. Tops up any underpayment from the 1st IRP6. File on eFiling before payment.",
            "",
        ]
    )
    annual.append(
        [
            "CIPC Annual Return",
            "~Aug 2026",
            "CIPC",
            "Upcoming",
            "Due within 30 business days of July incorporation anniversary. Fee: R100 (turnover under R500k). Late filing triggers 50% surcharge and deregistration proceedings. File at https://www.cipc.co.za",
            "",
        ]
    )
    annual.append(
        [
            "Income Tax Return - ITR14",
            "30 Jun 2027",
            "SARS",
            "Upcoming",
            "Covers FY 1 Jul 2025 - 30 Jun 2026. Due 12 months after year-end. Attach financial statements. File via SARS eFiling.",
            "",
        ]
    )

    vat = wb.create_sheet("VAT")
    vat.append(["tag", "text"])
    vat.append(["NOT APPLICABLE", "Expected rental income is under R500k per annum, well below the R1M compulsory VAT registration threshold. Voluntary registration is not recommended at this stage - it adds compliance cost with no material benefit."])

    employment = wb.create_sheet("Employment")
    employment.append(["obligation", "applicable", "trigger"])
    employment.append(["PAYE / EMP201 (monthly)", "No", "First direct employee"])
    employment.append(["UIF registration", "No", "First direct employee"])
    employment.append(["SDL (1% of payroll)", "No", "Payroll exceeds R500k/year"])
    employment.append(["COIDA (Return of Earnings by 30 Jun)", "No", "First direct employee"])

    employment_meta = wb.create_sheet("EmploymentMeta")
    employment_meta.append(["tag", "text"])
    employment_meta.append(["NOT APPLICABLE", "FPP has no employees. The managing agent is responsible for its own staff obligations. Reassess if FPP hires directly in future."])

    risks = wb.create_sheet("RiskFlags")
    risks.append(["number", "title", "description"])
    risks.append(["1", "1st IRP6 is overdue (31 Dec 2025)", "Priority action is to set up eFiling and file the late return. SARS can levy a 20% underestimation penalty plus interest on outstanding provisional tax."])
    risks.append(["2", "Short-term rental and mortgage", "Confirm with the bond provider that the mortgage agreement permits Airbnb-style short-term letting. Some South African lenders restrict this - a breach could trigger a material adverse clause."])
    risks.append(["3", "Managing agent relationship", "Confirm the agent is not structured in a way that creates an employment relationship with FPP. If SARS deems it one, PAYE obligations apply retroactively."])
    risks.append(["4", "Shareholder loan accounts", "All shareholder contributions accrue interest at Prime + 2% compounded monthly (SHA cl. 9). These must be tracked monthly and constitute taxable interest income in the hands of shareholders."])

    resources = wb.create_sheet("Resources")
    resources.append(["name", "url", "label"])
    resources.append(["SARS eFiling", "https://www.sars.gov.za/efiling", "sars.gov.za/efiling ->"])
    resources.append(["SARS Provisional Tax Guide", "https://www.sars.gov.za/types-of-tax/provisional-tax/", "sars.gov.za ->"])
    resources.append(["SARS Corporate Tax (ITR14)", "https://www.sars.gov.za/businesses-and-employers/companies/", "sars.gov.za ->"])
    resources.append(["CIPC Annual Returns", "https://www.cipc.co.za/index.php/manage-your-business/annual-returns/", "cipc.co.za ->"])
    resources.append(["CIPC Beneficial Ownership", "https://www.cipc.co.za/index.php/manage-your-business/beneficial-ownership/", "cipc.co.za ->"])
    resources.append(["POPIA - Information Regulator", "https://www.inforegulator.org.za", "inforegulator.org.za ->"])
    resources.append(["Companies Act 71 of 2008", "https://www.gov.za/documents/companies-act", "gov.za ->"])

    wb.save(path)


def sheet_rows(ws):
    return [
        [cell if cell is not None else "" for cell in row]
        for row in ws.iter_rows(min_row=2, values_only=True)
        if any(cell not in (None, "") for cell in row)
    ]


def generate_html() -> str:
    wb = load_workbook(TRACKER_PATH, data_only=True)
    meta_map = {k: v for k, v in sheet_rows(wb["Meta"])}

    once_off_rows = sheet_rows(wb["OnceOff"])
    annual_rows = sheet_rows(wb["Annual"])
    vat_row = sheet_rows(wb["VAT"])[0]
    emp_meta_row = sheet_rows(wb["EmploymentMeta"])[0]
    employment_rows = sheet_rows(wb["Employment"])
    risk_rows = sheet_rows(wb["RiskFlags"])
    resource_rows = sheet_rows(wb["Resources"])

    once_off_html = []
    for obligation, authority, status, notes in once_off_rows:
        once_off_html.append(
            f"""
        <tr>
          <td>{escape(str(obligation))}</td>
          <td><span class="auth {auth_class(str(authority))}">{escape(str(authority))}</span></td>
          <td><span class="status {status_class(str(status))}">{escape(str(status))}</span></td>
          <td class="note-text">{linkify(str(notes))}</td>
        </tr>"""
        )

    annual_html = []
    for obligation, due_date, authority, status, notes, overdue_row in annual_rows:
        row_class = ' class="overdue-row"' if str(overdue_row).strip().lower() in {"y", "yes", "true", "1"} else ""
        due_class = ' class="due-overdue"' if str(status).strip().lower() == "overdue" else ""
        annual_html.append(
            f"""
        <tr{row_class}>
          <td>{escape(str(obligation))}</td>
          <td{due_class}>{escape(str(due_date))}</td>
          <td><span class="auth {auth_class(str(authority))}">{escape(str(authority))}</span></td>
          <td><span class="status {status_class(str(status))}">{escape(str(status))}</span></td>
          <td class="note-text">{linkify(str(notes))}</td>
        </tr>"""
        )

    employment_html = []
    for obligation, applicable, trigger in employment_rows:
        employment_html.append(
            f"""
          <tr>
            <td>{escape(str(obligation))}</td>
            <td><span class="status {status_class(str(applicable))}">{escape(str(applicable))}</span></td>
            <td class="note-text">{escape(str(trigger))}</td>
          </tr>"""
        )

    risk_html = []
    for number, title, description in risk_rows:
        risk_html.append(
            f"""
      <div class="risk-card">
        <div class="risk-num">{escape(str(number))}</div>
        <div>
          <div class="risk-title">{escape(str(title))}</div>
          <div class="risk-desc">{escape(str(description))}</div>
        </div>
      </div>"""
        )

    resources_html = []
    for name, url, label in resource_rows:
        link_label = str(label).strip() or str(url)
        resources_html.append(
            f"""
      <div class="resource-card">
        <span class="resource-name">{escape(str(name))}</span>
        <a href="{escape(str(url))}" target="_blank" class="resource-link">{escape(link_label)}</a>
      </div>"""
        )

    year = datetime.now().year
    return f"""<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>FPP - Compliance Obligations</title>
<link rel="stylesheet" href="styles.css">
</head>
<body>
<div class="top-bar"></div>
<div class="page">
  <header class="header card">
    <h1>{escape(str(meta_map.get("title", "FPP Compliance Obligations")))}</h1>
    <div class="header-sub">
      <span><strong>{escape(str(meta_map.get("company", "")))}</strong></span>
      <span>Reg: {escape(str(meta_map.get("registration", "")))}</span>
      <span>Year-end: {escape(str(meta_map.get("year_end", "")))}</span>
      <span>As at {escape(str(meta_map.get("as_at", "")))}</span>
    </div>
    <div class="header-accent"></div>
  </header>

  <section class="summary card">
    {escape(str(meta_map.get("summary", "")))}
  </section>

  <section class="section card">
    <div class="section-header">
      <div class="section-icon icon-red">⚡</div>
      <h2>Once-Off Actions - Action Required Now</h2>
    </div>
    <table>
      <thead>
        <tr>
          <th class="col-obligation-wide">Obligation</th>
          <th class="col-authority">Authority</th>
          <th class="col-status">Status</th>
          <th>Notes</th>
        </tr>
      </thead>
      <tbody>{"".join(once_off_html)}
      </tbody>
    </table>
  </section>

  <section class="section card">
    <div class="section-header">
      <div class="section-icon icon-blue">📅</div>
      <h2>Annual Obligations</h2>
    </div>
    <table>
      <thead>
        <tr>
          <th class="col-obligation">Obligation</th>
          <th class="col-due">Due Date</th>
          <th class="col-authority-tight">Authority</th>
          <th class="col-status-tight">Status</th>
          <th>Notes</th>
        </tr>
      </thead>
      <tbody>{"".join(annual_html)}
      </tbody>
    </table>
  </section>

  <section class="section card">
    <div class="section-header">
      <div class="section-icon icon-green">🧾</div>
      <h2>VAT</h2>
    </div>
    <div class="info-box">
      <div class="na-tag">{escape(str(vat_row[0]))}</div>
      <p>{escape(str(vat_row[1]))}</p>
    </div>
  </section>

  <section class="section card">
    <div class="section-header">
      <div class="section-icon icon-green">👤</div>
      <h2>Employment-Related Obligations</h2>
    </div>
    <div class="info-box">
      <div class="na-tag">{escape(str(emp_meta_row[0]))}</div>
      <p>{escape(str(emp_meta_row[1]))}</p>
      <table class="emp-table">
        <thead>
          <tr>
            <th class="col-obligation-medium">Obligation</th>
            <th class="col-applicable">Applicable?</th>
            <th>Trigger</th>
          </tr>
        </thead>
        <tbody>{"".join(employment_html)}
        </tbody>
      </table>
    </div>
  </section>

  <section class="section card">
    <div class="section-header">
      <div class="section-icon icon-red">⚠️</div>
      <h2>Key Risk Flags</h2>
    </div>
    <div class="risk-list">{"".join(risk_html)}
    </div>
  </section>

  <section class="section card">
    <div class="section-header">
      <div class="section-icon icon-purple">🔗</div>
      <h2>Relevant Legislation &amp; Resources</h2>
    </div>
    <div class="resources-grid">{"".join(resources_html)}
    </div>
  </section>

  <footer class="footer">
    {escape(str(meta_map.get("footer", "")))}
    <div>Generated from compliance_tracker.xlsx ({year}).</div>
  </footer>
</div>
</body>
</html>
"""


def main() -> None:
    if not TRACKER_PATH.exists():
        create_default_tracker(TRACKER_PATH)
    html = generate_html()
    OUTPUT_HTML.write_text(html, encoding="utf-8")
    print(f"Wrote {OUTPUT_HTML.name} from {TRACKER_PATH.name}")


if __name__ == "__main__":
    main()
